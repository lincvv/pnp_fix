import PySimpleGUI as sg
from pathlib import Path
import openpyxl
import csv
import os

ROTATION_270 = 270
ROTATION_90 = 90
ROTATION_180 = 180
ROTATION_0 = 0
# Make .exe
# python -m pysimplegui-exemaker.pysimplegui-exemaker


class FileConverter:
    def __init__(self):
        self.SUFFIX_XLSX = ".xlsx"
        self.SUFFIX_CSV = ".csv"
        self.csv_file = None
        self.excel_file = None

    def csv_to_excel(self, csv_file):
        print("to excel")
        csv_data = []
        with open(csv_file) as file:
            reader = csv.reader(file)
            for row in reader:
                csv_data.append(row)
        wb = openpyxl.Workbook()
        sheet = wb.active
        for row in csv_data:
            sheet.append(row)
        self.excel_file = self.get_output_filename(csv_file, self.SUFFIX_XLSX)
        wb.save(self.excel_file)
        wb.close()

    def excel_to_csv(self, excel_file):
        print("to csv")
        wb = openpyxl.load_workbook(filename=excel_file)
        ws = wb.active
        self.csv_file = self.get_output_filename(excel_file, self.SUFFIX_CSV)
        with open(self.csv_file, 'w', newline="") as file:
            writer = csv.writer(file)
            for line in ws.iter_rows():
                writer.writerow([cell.value for cell in line])
        wb.close()

    def convert_file(self, input_file):
        if Path(input_file).suffix == self.SUFFIX_XLSX:
            self.excel_file = input_file
            self.excel_to_csv(input_file)
        elif Path(input_file).suffix == self.SUFFIX_CSV:
            self.csv_file = input_file
            self.csv_to_excel(input_file)
        else:
            raise TypeError("Only .csv or .xlsx files")

    def get_output_filename(self, input_file, sufix):
        return os.path.splitext(input_file)[0] + sufix


class DataManipulator:
    def __init__(self):
        self.CONVERTED_FACTOR = 0.0254
        self.NEGATIVE_SIGN = '-'
        self.MAX_ROTATION = 360

    def read_col(self, sh, col_name):
        list_data_col = list()
        for col_cell in sh.iter_cols(1, sh.max_column):
            if col_cell[0].value and col_cell[0].value.casefold() == col_name.casefold():
                # print("{} - {}".format(type(col_cell[0].value), col_cell[0]))
                for cell_data in col_cell[1:]:
                    list_data_col.append(cell_data.value)
        # print(list_data_col)
        return list_data_col

    def change_data(self, sh, col_name, list_val, prefix=None):
        for col_cell in sh.iter_cols(1, sh.max_column):  # iterate column cell
            if col_cell[0].value and col_cell[0].value.casefold() == col_name.casefold():
                for it in range(len(list_val)):
                    if prefix:
                        if list_val[it][0] == self.NEGATIVE_SIGN:
                            col_cell[it + 1].value = list_val[it][1:]
                        else:
                            col_cell[it + 1].value = prefix + list_val[it]
                    else:
                        col_cell[it + 1].value = list_val[it]
                break

    def convert_to_mm(self, name, column_data):
        if column_data[0].value and column_data[0].value.casefold() == name.casefold():
            for data_cell in column_data[1:]:
                try:
                    data_cell.value = format(float(data_cell.value) * self.CONVERTED_FACTOR, ".2f")
                except TypeError:
                    data_cell.value = None
                except ValueError:
                    raise ValueError(f"Non-numeric value {data_cell.value} cannot be converted to float.")

    def calculate_new_rot_value(self, value):
        return 0 if float(value) == ROTATION_270 or float(value) > self.MAX_ROTATION else float(value) + ROTATION_90

    def strip_negative_sign(self, pos_list):
        return [item[1:] if item[0] == self.NEGATIVE_SIGN else item for item in pos_list]


def print_text(text_element, text_data):
    text = text_element
    text.update(text_data)


def main():
    file_converter = FileConverter()
    data_man = DataManipulator()
    name_pos_x = "X"
    name_pos_y = "Y"
    name_rot = 'R'
    rotated = ROTATION_0
    wb, sheet = None, None
    is_init = False

    sg.theme("DarkGrey9")
    sg.set_options(font=("Microsoft JhengHei", 16))

    layout_title = [
        [sg.Input(enable_events=True, key='-INPUT-', size=(25, 1)),
         sg.FileBrowse(key='-FILE-', file_types=(("CSV Files", "*.csv"), ("XLSX", ".xlsx"))),
         ],
        [sg.Button("Flip-X", enable_events=True, key='-FLIP-X-', font='Helvetica 16', pad=(45, 10),),
         sg.Button("Flip-Y", enable_events=True, key='-FLIP-Y-', font='Helvetica 16', pad=(15, 0),),
         sg.Checkbox("to mm", key="-CHECKBOX-", default=False, enable_events=True, metadata=False, size=(35, 1))
         ],

        [sg.Text("Rotated: {}°".format(rotated), size=(35, 0), key='-text-', font='Helvetica 16', pad=(95, 5))],
        [sg.Button('', enable_events=True, key='-ROTATE-', font='Helvetica 16', image_filename='arrow.png',
                   pad=(110, 0), button_color=('white', 'white')),
         sg.Button('', enable_events=True, key='-SAVE-', font='Helvetica 16', image_filename='save.png',
                   pad=(10, 0), button_color=('white', 'white')),
         ],
        [sg.Text("", size=(50, 10), key='-error-', font='Arial 12', pad=(5, 5), text_color="red")],
    ]

    window = sg.Window('FixPnP', layout_title, resizable=False, icon="01_nti_blue.ico", auto_size_buttons=True,
                       size=(450, 260))

    while True:
        event, values = window.read()
        if event == '-INPUT-':
            input_file = values['-INPUT-']
            if Path(input_file).is_file():
                try:
                    file_converter.convert_file(input_file)
                except (FileNotFoundError, TypeError) as e:
                    print_text(window['-error-'], e)
                wb = openpyxl.load_workbook(file_converter.excel_file)
                sheet = wb.active
                is_init = True
                print_text(window['-error-'], "")
                print_text(window['-text-'], "Rotated: {}°".format(ROTATION_0))
                rotated = ROTATION_0

        elif event == "-CHECKBOX-" and values["-CHECKBOX-"] and is_init:
            try:
                for column_cell in sheet.iter_cols(1, sheet.max_column):
                    data_man.convert_to_mm(name_pos_x, column_cell)
                    data_man.convert_to_mm(name_pos_y, column_cell)
            except Exception as e:
                # print("Error: ", e)
                print_text(window['-error-'], "Error: {}".format(e))

        elif event == '-FLIP-X-' and is_init:
            try:
                data_man.change_data(sheet, name_pos_x, data_man.read_col(sheet, name_pos_x),
                                     prefix=data_man.NEGATIVE_SIGN)
            except Exception as e:
                # print("Error: ", e)
                print_text(window['-error-'], "Error: {}".format(e))

        elif event == '-FLIP-Y-' and is_init:
            try:
                data_man.change_data(sheet, name_pos_y, data_man.read_col(sheet, name_pos_y),
                                     prefix=data_man.NEGATIVE_SIGN)
            except Exception as e:
                # print("Error: ", e)
                print_text(window['-error-'], "Error: {}".format(e))

        elif event == '-ROTATE-' and is_init:
            rotated += ROTATION_90
            if rotated > ROTATION_270:
                rotated = ROTATION_0

            print_text(window['-text-'], "Rotated: {}°".format(rotated))

            try:
                for column_cell in sheet.iter_cols(1, sheet.max_column):
                    if column_cell[0].value and column_cell[0].value.casefold() == name_rot.casefold():
                        for data in column_cell[1:]:
                            data.value = data_man.calculate_new_rot_value(data.value)
                        break

                list_pos_x = data_man.read_col(sheet, name_pos_x)
                list_pos_y = data_man.read_col(sheet, name_pos_y)
                data_man.change_data(sheet, name_pos_x, list_pos_y, prefix=data_man.NEGATIVE_SIGN)
                data_man.change_data(sheet, name_pos_y, list_pos_x)
            except Exception as e:
                # print("Error: ", e)
                print_text(window['-error-'], "Error: {}".format(e))

        elif event == '-SAVE-':
            if wb and Path(file_converter.excel_file).is_file():
                try:
                    wb.save(file_converter.excel_file)
                    file_converter.excel_to_csv(file_converter.excel_file)
                except Exception as e:
                    print_text(window['-error-'], "Error: {}".format(e))
            else:
                print_text(window['-error-'], "Error: Wrong file type")

        elif event in (sg.WIN_CLOSED, 'Exit'):
            if wb:
                wb.close()
            break

    window.close()


if __name__ == '__main__':
    main()
