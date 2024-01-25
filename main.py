import PySimpleGUI as sg
from pathlib import Path
import openpyxl
import csv
import os

NEGATIVE_SIGN = '-'
ROTATION_270 = 270
ROTATION_90 = 90
ROTATION_180 = 180
ROTATION_0 = 0
CONVERTED_FACTOR = 0.0254


class FileConverter:
    def __init__(self):
        self.SUFIX_XLSX = ".xlsx"
        self.SUFIX_CSV = ".csv"

    def csv_to_excel(self, csv_file, excel_file):
        print("to excel")
        csv_data = []
        with open(csv_file) as file:
            reader = csv.reader(file)
            for row in reader:
                csv_data.append(row)
        wb = openpyxl.Workbook()
        sheet = wb.active
        for row in csv_data:
            sheet.append(row)
        wb.save(excel_file)
        wb.close()

    def excel_to_csv(self, excel_file, csv_file):
        print("to csv")
        wb = openpyxl.load_workbook(filename=excel_file)
        ws = wb.active
        with open(csv_file, 'w', newline="") as file:
            writer = csv.writer(file)
            for line in ws.iter_rows():
                writer.writerow([cell.value for cell in line])
        wb.close()

    def create_file(self, input_file, sufix ):
        output_file = os.path.splitext(input_file)[0] + sufix
        # if not Path(output_file).is_file():
        print("path")
        if sufix == self.SUFIX_XLSX:
            self.csv_to_excel(input_file, output_file)
        if sufix == self.SUFIX_CSV:
            self. excel_to_csv(input_file, output_file)
        return output_file


class DataManipulator:
    def __init__(self):
        self.CONVERTED_FACTOR = 0.0254
        self.NEGATIVE_SIGN = '-'
        self.MAX_ROTATION = 360

    def read_col(self, sh, col_name):
        list_data_col = list()
        for col_cell in sh.iter_cols(1, sh.max_column):
            if col_cell[0].value == col_name:
                for cell_data in col_cell[1:]:
                    list_data_col.append(cell_data.value)
        print(list_data_col)
        return list_data_col

    def change_data(self, sh, col_name, list_val, prefix=None):
        for col_cell in sh.iter_cols(1, sh.max_column):  # iterate column cell
            if col_cell[0].value == col_name:
                for it in range(len(list_val)):
                    if prefix:
                        if list_val[it][0] == NEGATIVE_SIGN:
                            col_cell[it + 1].value = list_val[it][1:]
                        else:
                            col_cell[it + 1].value = prefix + list_val[it]
                    else:
                        col_cell[it + 1].value = list_val[it]
                break

    def convert_to_mm(self, name, column_data):
        if column_data[0].value == name:
            for data_cell in column_data[1:]:
                print(format(float(data_cell.value), ".2f",))
                try:
                    data_cell.value = format(float(data_cell.value) * self.CONVERTED_FACTOR, ".2f")
                except TypeError:
                    data_cell.value = None
                except ValueError:
                    print(f"Non-numeric value {data_cell.value} cannot be converted to float.")

    def calculate_new_rot_value(self, value):
        return 0 if float(value) == ROTATION_270 or float(value) > self.MAX_ROTATION else float(value) + ROTATION_90

    def strip_negative_sign(self, pos_list):
        return [item[1:] if item[0] == self.NEGATIVE_SIGN else item for item in pos_list]


def print_text(text_element, text_data):
    text = text_element
    text.update(text_data)


def main():
    file_converter = FileConverter()
    data_man = DataManipulator()
    name_pos_x = "x"
    name_pos_y = "y"
    name_rot = 'r'
    rotated = ROTATION_0
    excel_file, csv_file, wb, sheet = None, None, None, None
    is_init = False

    sg.theme("DarkGrey9")
    sg.set_options(font=("Microsoft JhengHei", 16))

    layout_title = [
        [sg.Input(enable_events=True, key='-INPUT-', size=(25, 1)),
         sg.FileBrowse(key='-FILE-', file_types=(("CSV Files", "*.csv"), ("XLSX", ".xlsx"))),
         ],
        [sg.Button("Flip-X", enable_events=True, key='-FLIP-X-', font='Helvetica 16', pad=(45, 10),),
         sg.Button("Flip-Y", enable_events=True, key='-FLIP-Y-', font='Helvetica 16', pad=(15, 0),),
         sg.Checkbox("to mm", key="-CHECKBOX-", default=False, enable_events=True, metadata=False, size=(35, 1))
         ],

        [sg.Text("Rotated: {}°".format(rotated), size=(35, 0), key='-text-', font='Helvetica 16', pad=(95, 5))],
        [sg.Button('', enable_events=True, key='-ROTATE-', font='Helvetica 16', image_filename='arrow.png',
                   pad=(110, 0), button_color=('white', 'white')),
         sg.Button('', enable_events=True, key='-SAVE-', font='Helvetica 16', image_filename='save.png',
                   pad=(10, 0), button_color=('white', 'white')),
         ],
        [sg.Text("", size=(50, 10), key='-error-', font='Arial 12', pad=(5, 5), text_color="red")],
    ]

    window = sg.Window('Fix PnP', layout_title, resizable=False, icon="01_nti_blue.ico", auto_size_buttons=True,
                       size=(450, 260))

    while True:
        event, values = window.read()
        if event == '-INPUT-':
            input_file = values['-INPUT-']
            if Path(input_file).is_file():
                try:
                    if Path(input_file).suffix == ".csv":
                        csv_file = input_file
                        excel_file = file_converter.create_file(input_file, file_converter.SUFIX_XLSX)
                    elif Path(input_file).suffix == ".xlsx":
                        excel_file = input_file
                        csv_file = file_converter. create_file(input_file, file_converter.SUFIX_CSV)
                    else:
                        print_text(window['-error-'], "Only .csv or .xlsx files")
                        continue
                except FileNotFoundError as e:
                    print_text(window['-error-'], e)
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active
                is_init = True
                print_text(window['-error-'], "")
                print_text(window['-text-'], "Rotated: {}°".format(ROTATION_0))
                rotated = ROTATION_0

        elif event == "-CHECKBOX-" and values["-CHECKBOX-"] and is_init:
            try:
                for column_cell in sheet.iter_cols(1, sheet.max_column):
                    data_man.convert_to_mm(name_pos_x, column_cell)
                    data_man.convert_to_mm(name_pos_y, column_cell)
            except Exception as e:
                print("Error: ", e)

        elif event == '-FLIP-X-' and is_init:
            try:
                data_man.change_data(sheet, name_pos_x, data_man.read_col(sheet, name_pos_x),
                                     prefix=data_man.NEGATIVE_SIGN)
            except Exception as e:
                print("Error: ", e)

        elif event == '-FLIP-Y-' and is_init:
            try:
                data_man.change_data(sheet, name_pos_y, data_man.read_col(sheet, name_pos_y),
                                     prefix=data_man.NEGATIVE_SIGN)
            except Exception as e:
                print("Error: ", e)

        elif event == '-ROTATE-' and is_init:
            rotated += ROTATION_90
            if rotated > ROTATION_270:
                rotated = ROTATION_0

            print_text(window['-text-'], "Rotated: {}°".format(rotated))

            try:
                for column_cell in sheet.iter_cols(1, sheet.max_column):
                    if column_cell[0].value == name_rot:
                        for data in column_cell[1:]:
                            data.value = data_man.calculate_new_rot_value(data.value)
                        break

                list_pos_x = data_man.read_col(sheet, name_pos_x)
                list_pos_y = data_man.read_col(sheet, name_pos_y)
                data_man.change_data(sheet, name_pos_x, list_pos_y, prefix='-')
                data_man.change_data(sheet, name_pos_y, list_pos_x)
            except Exception as e:
                print("Error: ", e)

        elif event == '-SAVE-':
            if wb:
                wb.save(excel_file)
            try:
                file_converter.excel_to_csv(csv_file=csv_file, excel_file=excel_file)
            except TypeError:
                print_text(window['-error-'], "Error: Wrong file type")

        elif event in (sg.WIN_CLOSED, 'Exit'):
            if wb:
                wb.close()
            break

    window.close()


if __name__ == '__main__':
    main()